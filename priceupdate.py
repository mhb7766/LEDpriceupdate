#some changes
import openpyxl
from openpyxl import Workbook
import os
from tkinter import *
from tkinter.filedialog import askopenfilename
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

#set markup
markup = input("Enter markup amount (e.g. 1.33): ")
#set price column
col_alpha = input("Enter alphabetical column where price resides (e.g. C): ")
col = column_index_from_string(col_alpha)

#choose pricesheet
input("Hit enter to choose price sheet ")
Tk().withdraw()
pricesheet=askopenfilename()

#choose current inventory sheet
input("Hit enter to choose sheet containing current listings on our site ")
currentlistings=askopenfilename()

#open workbook
pr = openpyxl.load_workbook(pricesheet)
price = pr['Sheet1']
li = openpyxl.load_workbook(currentlistings)
listings = li['onsite']

#create new workbook for updated data and name columns
new = Workbook()
sheet = new.active
sheet.cell(row=1, column=1, value='productcode')
sheet.cell(row=1, column=2, value='productprice')
sheet.cell(row=1, column=3, value='saleprice')

#change directory
#os.chdir('c:\\users\\mike\\pictures\\led')

################## ONE WORKBOOK METHOD #################
#set variable for sheets 
#onsite = wb['onsite']
#allmax = wb['allmax']
########################################################

#add markup to price and store in new worksheet
#emptycount tracks 0 prices and subtracts them from the index to avoid blank records
emptycount=0
for i in range (2, price.max_row):
	updateprice = float(price.cell(row=i, column=col).value)*1.33
	if updateprice > 0:
		sheet.cell(row=i-emptycount, column=1, value=price.cell(row=i, column=1).value)
		sheet.cell(row=i-emptycount, column=2,value=str(round(updateprice,2)))
	else: emptycount+=1

#need to make sure file is saved as .csv
#savelocation = input("Hit enter to choose a save location for the output file ")
new.save('readyforupload.xlsx')